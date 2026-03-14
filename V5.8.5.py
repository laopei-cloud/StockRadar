# ==================================================================================
# 主力行为实时监测雷达 V5.8.5 - 共振连续榜计算修复版
# 版本特点：量额比(VPR) | 拉升信心 | 完整续跑 | 涨跌停精准识别(增强版) | 共振连续5日榜 
#          | 强庄👑标记 | 热力条 | 频率统计 | 价格缓存 | 组合策略 | 智能对齐 | 终端自适应
#          | 博弈比负偏取整修复 | 委比分语义优化 | 量额比方向修正 | 共振榜意图修复
#          | 昨收盘价多重获取 | 涨跌停互斥判断 | 板块差异识别 | 连续天数计算修复
# 修复日期：2025-02-13
# 核心修复：共振连续天数基于实际连续日期计算，而非历史总长度
# ==================================================================================

import pandas as pd
import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from pytdx.hq import TdxHq_API
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from io import BytesIO
import datetime
import os
import time
import warnings
from tqdm import tqdm
from matplotlib.figure import Figure
import h5py
import json
import sys
from colorama import init, Fore, Back, Style
import concurrent.futures
from threading import Lock, Thread
import queue
import threading
import re
from collections import deque

# 初始化colorama
init(autoreset=True)

# --- 1. 基础配置与环境 ---
warnings.filterwarnings("ignore")
plt.rcParams['font.sans-serif'] = ['SimHei']
plt.rcParams['axes.unicode_minus'] = False

# --- 全局意图计数器 ---
intent_counter = {}

# --- 全局报警统计 ---
warning_stats = {
    'total_warnings': 0,
    'warning_stocks': set(),
    'warning_by_type': {}
}

# --- 性能统计 ---
performance_stats = {
    'total_scans': 0,
    'total_stocks_processed': 0,
    'h5_write_success': 0,
    'h5_write_failed': 0,
    'h5_queue_full_count': 0,
    'start_time': time.time()
}

# --- 股票数据缓存（用于跨函数传递最新价和状态）---
stock_data_cache = {}
stock_data_cache_lock = Lock()

# --- 委比历史平滑 ---
imbalance_history = {}
imbalance_history_lock = Lock()

# --- 分数历史趋势 ---
score_history = {}
score_history_lock = Lock()

# --- 核心共振连续上榜统计（5日周期）---
resonance_history = {}  # {code: deque([date1, date2, ...])} 按时间顺序存储
resonance_streak = {}   # {code: 连续天数} - ✅ 修复：基于实际连续日期计算
resonance_history_lock = Lock()

RESONANCE_CONFIG = {
    'streak_days': 5,                    # 统计周期
    'display_top': 10,                   # 显示前几名
    'streak_colors': {
        1: Fore.WHITE,                    # 1天
        2: Fore.CYAN,                      # 2天
        3: Fore.YELLOW,                     # 3天
        4: Fore.MAGENTA,                      # 4天
        5: Fore.RED + Style.BRIGHT             # 5天
    }
}

# --- 多信号组合矩阵策略配置 ---
SIGNAL_COMBINATION_CONFIG = {
    'combinations': {
        'resonance_top1': {
            'name': '共振+Top1',
            'conditions': ['is_resonance', 'is_top1'],
            'confidence': '⭐⭐⭐⭐⭐',
            'strategy': '重点配置',
            'color': Fore.RED + Style.BRIGHT,
            'alert': True
        },
        'resonance_vpr_strong': {
            'name': '共振+溢价>1%',
            'conditions': ['is_resonance', 'vpr_strong_buy'],
            'confidence': '⭐⭐⭐⭐⭐',
            'strategy': '积极跟随',
            'color': Fore.RED + Style.BRIGHT,
            'alert': True
        },
        'strong_new_top10': {
            'name': '强势+新晋Top10',
            'conditions': ['is_strong', 'is_new_top10'],
            'confidence': '⭐⭐⭐⭐',
            'strategy': '轻仓试多',
            'color': Fore.YELLOW,
            'alert': True
        },
        'warning_vpr_weak': {
            'name': '报警+折价>1%',
            'conditions': ['is_warning', 'vpr_strong_sell'],
            'confidence': '⭐⭐⭐⭐',
            'strategy': '坚决卖出',
            'color': Fore.MAGENTA,
            'alert': True
        },
        'resonance_streak_3': {
            'name': '共振+连续3天',
            'conditions': ['is_resonance', 'streak_3'],
            'confidence': '⭐⭐⭐⭐⭐',
            'strategy': '强庄确认',
            'color': Fore.RED + Style.BRIGHT,
            'alert': True
        },
        'resonance_streak_5': {
            'name': '共振+连续5天',
            'conditions': ['is_resonance', 'streak_5'],
            'confidence': '⭐⭐⭐⭐⭐⭐',
            'strategy': '绝对龙头',
            'color': Fore.RED + Back.YELLOW + Style.BRIGHT,
            'alert': True
        },
        'strong_vpr_mild': {
            'name': '强势+溢价>0.5%',
            'conditions': ['is_strong', 'vpr_mild_buy'],
            'confidence': '⭐⭐⭐',
            'strategy': '观察池',
            'color': Fore.YELLOW,
            'alert': False
        },
        'warning_z_bias_severe': {
            'name': '报警+Z偏<-0.2',
            'conditions': ['is_warning', 'z_bias_severe_sell'],
            'confidence': '⭐⭐⭐⭐',
            'strategy': '紧急卖出',
            'color': Fore.MAGENTA + Style.BRIGHT,
            'alert': True
        }
    },
    'display_top': 10,
    'enable_sound': False
}

# --- 组合信号历史（防重复报警）---
combination_history = {}
combination_history_lock = Lock()

# --- 扫描状态锁 ---
scan_in_progress = False
scan_lock = Lock()

# --- 线程安全锁 ---
intent_counter_lock = Lock()
warning_stats_lock = Lock()
h5_write_lock = Lock()
performance_lock = Lock()

# --- 生产者消费者队列 ---
h5_write_queue = queue.Queue(maxsize=500)

# --- 配置参数 - 涨跌停判断增强版 ---
CONFIG = {
    'display_mode': 'compact',
    'show_all_stocks': False,
    'min_z_bias': 0.0,
    'scan_interval': 30,
    'save_image': False,
    'h5_filename': 'all_stocks_data.h5',
    'enable_warning': True,
    'enable_sound': False,
    'max_workers': 3,
    'write_batch_size': 20,
    'enable_parallel': False,
    'queue_workers': {
        'h5_writer': 1,
    },
    'show_normal_signals': False,
    'min_display_score': 30,
    # 信号显示阈值
    'signal_display_threshold': {
        'resonance': 65,
        'warning': 35,
        'strong': 55,
        'positive': 45
    },
    'resonance_z_bias_threshold': 0.08,
    'resonance_score_threshold': 65,
    'warning_score_threshold': 35,
    'warning_z_bias_threshold': -0.15,
    'strong_z_bias_threshold': 0.05,
    'strong_score_threshold': 55,
    'positive_score_threshold': 45,
    'display_width': 120,
    'show_stock_name': True,
    'max_name_length': 8,
    'debug_mode': False,
    # 委比平滑配置
    'enable_order_imbalance': True,
    'imbalance_smooth_window': 3,
    'imbalance_strong_buy': 0.5,
    'imbalance_strong_sell': -0.5,
    'imbalance_buy_score': 10,
    'imbalance_sell_score': 15,
    # Z偏时间补偿
    'enable_z_bias_time_compensation': True,
    'z_bias_afternoon_multiplier': 1.3,
    # 量额比配置
    'enable_volume_price_ratio': True,
    'vpr_strong_buy': 1.01,
    'vpr_strong_sell': 0.99,
    'vpr_mild_buy': 1.005,
    'vpr_mild_sell': 0.995,
    'vpr_buy_score': 15,
    'vpr_sell_score': 15,
    'vpr_display_threshold': 1.005,
    # 趋势判断阈值（价格自适应）
    'trend_threshold_low': 0.002,
    'trend_threshold_medium': 0.001,
    'trend_threshold_high': 0.0005,
    # Z偏严重阈值
    'z_bias_severe_buy': 0.2,
    'z_bias_severe_sell': -0.2,
}

# ==================== 显示优化配置 ====================

# 自动检测终端宽度
try:
    TERMINAL_WIDTH = os.get_terminal_size().columns
except:
    TERMINAL_WIDTH = 120

# 显示配置（智能适配）
DISPLAY_CONFIG = {
    'terminal_width': min(TERMINAL_WIDTH, 140),
    'compact_mode': TERMINAL_WIDTH < 100,
    'spacing': 1,
    'width': {
        'rank': 4,
        'code': 8,
        'name': 10,
        'trend': 2,
        'intent': 6,
        'count': 3,
        'score': 5,
        'z_bias': 7,
        'imbalance': 7,
        'vpr': 6,
        'price': 9,
        'streak': 4,
        'bar': 7,
        'combo_name': 22,
        'confidence': 10,
        'strategy': 10,
        'intent_full': 12,
        'strength': 7,
        'frequency': 6,
    }
}

# 根据紧凑模式调整列宽
if DISPLAY_CONFIG['compact_mode']:
    DISPLAY_CONFIG['width']['name'] = 8
    DISPLAY_CONFIG['width']['intent'] = 4
    DISPLAY_CONFIG['width']['score'] = 4
    DISPLAY_CONFIG['width']['z_bias'] = 6
    DISPLAY_CONFIG['width']['imbalance'] = 6
    DISPLAY_CONFIG['width']['vpr'] = 5
    DISPLAY_CONFIG['width']['price'] = 8
    DISPLAY_CONFIG['width']['streak'] = 2
    DISPLAY_CONFIG['width']['bar'] = 0
    DISPLAY_CONFIG['width']['combo_name'] = 18
    DISPLAY_CONFIG['spacing'] = 0

# ==================== 显示优化工具函数 ====================

def get_visual_width(text):
    """计算字符串的视觉宽度（中文=2，英文/数字/符号=1）"""
    width = 0
    for ch in text:
        if '\u4e00' <= ch <= '\u9fff':
            width += 2
        else:
            width += 1
    return width

def pad_visual(text, target_width):
    """按视觉宽度填充空格到目标宽度"""
    current_width = get_visual_width(text)
    if current_width >= target_width:
        return text
    return text + ' ' * (target_width - current_width)

def trim_visual(text, max_width):
    """按视觉宽度截断字符串（保留完整字符）"""
    if get_visual_width(text) <= max_width:
        return text

    result = ''
    width = 0
    for ch in text:
        ch_width = 2 if '\u4e00' <= ch <= '\u9fff' else 1
        if width + ch_width > max_width - 1:
            result += '…'
            break
        result += ch
        width += ch_width
    return result

# ==================== 分数计算核心函数 ====================

def calculate_z_score(z_bias):
    """
    计算博弈比分（修复负偏取整问题）
    正偏：z_bias × 120，向下取整
    负偏：-|z_bias| × 80，向下取整后加负号
    """
    z_bias_clamped = max(-0.3, min(0.3, z_bias))

    if z_bias_clamped > 0:
        return int(z_bias_clamped * 120)
    elif z_bias_clamped < 0:
        return -int(abs(z_bias_clamped) * 80)
    else:
        return 0

# ==================== 持久化状态管理 ====================

SESSION_STATE = {
    'today': '',
    'initialized': False,
    'scan_count': 0,
    'resonance_total': 0,
    'last_save_time': ''
}

def init_session_state():
    """初始化/加载今日会话状态"""
    global SESSION_STATE, intent_counter, score_history, warning_stats
    global resonance_history, resonance_streak, combination_history

    today = datetime.date.today().strftime('%Y-%m-%d')
    SESSION_STATE['today'] = today

    if os.path.exists(CONFIG['h5_filename']):
        try:
            with h5py.File(CONFIG['h5_filename'], 'r') as f:
                if 'session_state' in f and today in f['session_state']:
                    state_data = json.loads(f['session_state'][today][()])

                    SESSION_STATE['scan_count'] = state_data.get('scan_count', 0)
                    SESSION_STATE['resonance_total'] = state_data.get('resonance_total', 0)
                    SESSION_STATE['initialized'] = True

                    if 'intent_counter' in state_data:
                        intent_counter.clear()
                        intent_counter.update(state_data['intent_counter'])

                    if 'score_history' in state_data:
                        score_history.clear()
                        for code, scores in state_data['score_history'].items():
                            score_history[code] = deque(scores, maxlen=3)

                    if 'warning_stats' in state_data:
                        warning_stats['total_warnings'] = state_data['warning_stats'].get('total_warnings', 0)
                        warning_stats['warning_stocks'] = set(state_data['warning_stats'].get('warning_stocks', []))
                        warning_stats['warning_by_type'] = state_data['warning_stats'].get('warning_by_type', {})

                    if 'resonance_history' in state_data:
                        resonance_history.clear()
                        for code, dates in state_data['resonance_history'].items():
                            resonance_history[code] = deque(dates, maxlen=RESONANCE_CONFIG['streak_days'])

                    if 'resonance_streak' in state_data:
                        # ✅ 修复：不直接从历史加载连续天数，而是重新计算
                        resonance_streak.clear()
                        for code, dates in resonance_history.items():
                            calculate_resonance_streak(code, list(dates))

                    if 'combination_history' in state_data:
                        combination_history.clear()
                        combination_history.update(state_data['combination_history'])

                    streak_count = len([d for d in resonance_streak.values() if d > 0])
                    print(f"{Fore.GREEN}🔄 续跑模式 | 今日已扫描 {SESSION_STATE['scan_count']} 轮 | "
                          f"累计共振 {SESSION_STATE['resonance_total']} 次 | "
                          f"连续共振股 {streak_count} 只")
                    return
        except Exception as e:
            if CONFIG['debug_mode']:
                print(f"{Fore.YELLOW}⚠️ 加载会话状态失败: {e}")

    print(f"{Fore.CYAN}🆕 今日首次运行 | 初始化新会话")
    SESSION_STATE['initialized'] = True

def save_session_state():
    """保存当前会话状态到H5"""
    if not SESSION_STATE['initialized']:
        return

    try:
        today = SESSION_STATE['today']
        state_data = {
            'scan_count': SESSION_STATE['scan_count'],
            'resonance_total': SESSION_STATE['resonance_total'],
            'last_save_time': datetime.datetime.now().isoformat(),
            'intent_counter': {k: v for k, v in intent_counter.items()},
            'score_history': {code: list(scores) for code, scores in score_history.items()},
            'warning_stats': {
                'total_warnings': warning_stats['total_warnings'],
                'warning_stocks': list(warning_stats['warning_stocks']),
                'warning_by_type': warning_stats['warning_by_type']
            },
            'resonance_history': {code: list(dates) for code, dates in resonance_history.items()},
            'resonance_streak': dict(resonance_streak),
            'combination_history': dict(combination_history)
        }

        with h5py.File(CONFIG['h5_filename'], 'a') as f:
            if 'session_state' not in f:
                f.create_group('session_state')

            dt = h5py.special_dtype(vlen=str)
            if today in f['session_state']:
                del f['session_state'][today]

            f['session_state'].create_dataset(
                today,
                data=np.array(json.dumps(state_data, ensure_ascii=False), dtype=object),
                dtype=dt
            )
    except Exception as e:
        if CONFIG['debug_mode']:
            print(f"{Fore.YELLOW}⚠️ 保存会话状态失败: {e}")

# ==================== 股票名称管理 ====================

STOCK_NAME_DB = {}

def load_stock_names_from_tdx():
    """从通达信本地文件加载股票名称"""
    global STOCK_NAME_DB
    loaded_count = 0

    if os.path.exists('stocks.txt'):
        try:
            with open('stocks.txt', 'r', encoding='utf-8') as f:
                for line in f:
                    line = line.strip()
                    if not line:
                        continue

                    code_match = re.search(r'(\d{6})', line)
                    if code_match:
                        code = code_match.group(1).zfill(6)
                        name = None

                        if '\t' in line:
                            parts = line.split('\t')
                            if len(parts) >= 2:
                                name = parts[1].strip()
                        elif ' ' in line:
                            parts = line.split(' ')
                            if len(parts) >= 2:
                                name = parts[1].strip()
                        elif ',' in line:
                            parts = line.split(',')
                            if len(parts) >= 2:
                                name = parts[1].strip()
                        else:
                            name_part = line[6:].strip()
                            if name_part:
                                name = name_part

                        if name:
                            STOCK_NAME_DB[code] = name
                            loaded_count += 1
        except Exception as e:
            if CONFIG['debug_mode']:
                print(f"{Fore.YELLOW}⚠️ 读取stocks.txt失败: {e}")

    sector_files = ['sector_map.txt', 'sector.txt', 'industry.txt']
    for filename in sector_files:
        if os.path.exists(filename):
            try:
                for encoding in ['utf-8', 'gbk', 'gb2312', 'utf-8-sig']:
                    try:
                        with open(filename, 'r', encoding=encoding) as f:
                            for line in f:
                                line = line.strip()
                                if not line:
                                    continue

                                code_match = re.search(r'(\d{6})', line)
                                if not code_match:
                                    continue

                                code = code_match.group(1).zfill(6)
                                if code in STOCK_NAME_DB:
                                    continue

                                code_pos = line.find(code)
                                if code_pos >= 0:
                                    name_part = line[code_pos + 6:].strip()
                                    for sep in ['\t', ',', ' ', '|', '，', '　']:
                                        if name_part.startswith(sep):
                                            name_part = name_part[1:].strip()

                                    if name_part:
                                        STOCK_NAME_DB[code] = name_part
                                        loaded_count += 1
                            break
                    except UnicodeDecodeError:
                        continue
            except Exception:
                continue

    print(f"{Fore.GREEN}✅ 股票名称库加载完成: {loaded_count} 只")
    return loaded_count

def get_stock_name(code):
    """获取股票名称"""
    code = str(code).zfill(6)
    if code in STOCK_NAME_DB:
        return STOCK_NAME_DB[code]
    return f"{code}"

def get_latest_price(code):
    """获取股票最新价（从缓存或H5）"""
    with stock_data_cache_lock:
        if code in stock_data_cache:
            cache_time = stock_data_cache[code].get('time')
            if cache_time and (datetime.datetime.now() - cache_time).seconds < 60:
                return stock_data_cache[code].get('price', 0.0)

    today = datetime.date.today().strftime('%Y-%m-%d')
    try:
        with h5py.File(CONFIG['h5_filename'], 'r') as f:
            group_name = f"group_{code[:2]}"
            if group_name in f and code in f[group_name] and today in f[group_name][code]:
                indicators_data = f[group_name][code][today]['indicators'][()]
                if isinstance(indicators_data, bytes):
                    indicators = json.loads(indicators_data.decode('utf-8'))
                else:
                    indicators = json.loads(indicators_data)
                return indicators.get('price', 0.0)
    except:
        pass
    return 0.0

# ==================== 获取最新共振意图 ====================

def get_latest_resonance_intent(code):
    """
    获取股票最近一次共振时的意图
    用于共振连续榜显示，避免显示历史错误意图
    """
    # 1. 优先从缓存获取（今日共振状态）
    with stock_data_cache_lock:
        if code in stock_data_cache:
            data = stock_data_cache[code]
            # 只返回共振状态的意图
            if data.get('is_resonance', False):
                return data.get('intent', '⚖️均衡')

    # 2. 从H5获取今日共振记录
    today = datetime.date.today().strftime('%Y-%m-%d')
    try:
        with h5py.File(CONFIG['h5_filename'], 'r') as f:
            group_name = f"group_{code[:2]}"
            if group_name in f and code in f[group_name] and today in f[group_name][code]:
                indicators_data = f[group_name][code][today]['indicators'][()]
                if isinstance(indicators_data, bytes):
                    indicators = json.loads(indicators_data.decode('utf-8'))
                else:
                    indicators = json.loads(indicators_data)

                # 只取共振信号的意图
                if indicators.get('signal') == 'resonance':
                    return indicators.get('intent', '⚖️均衡')
    except:
        pass

    return '⚖️均衡'

# ==================== ✅ 新增：昨收盘价多重获取函数 ====================

def get_pre_close(api, market, code):
    """
    获取昨收盘价，多重保障
    1. 从实时行情获取 pre_close
    2. 从日线数据获取
    3. 从缓存获取
    4. 从H5历史数据获取
    """
    # 方法1：从实时行情获取
    try:
        quotes = api.get_security_quotes([(market, code)])
        if quotes and len(quotes) > 0:
            pre_close = quotes[0].get('pre_close', 0)
            if pre_close and float(pre_close) > 0:
                return float(pre_close)
    except:
        pass

    # 方法2：从日线数据获取
    try:
        bars = api.get_security_bars(9, market, code, 0, 1)  # 9=日线
        if bars and len(bars) > 0:
            return float(bars[0]['close'])  # 昨日收盘价
    except:
        pass

    # 方法3：从缓存获取
    with stock_data_cache_lock:
        if code in stock_data_cache and 'pre_close' in stock_data_cache[code]:
            return stock_data_cache[code]['pre_close']

    # 方法4：从H5获取昨日数据
    yesterday = (datetime.date.today() - datetime.timedelta(days=1)).strftime('%Y-%m-%d')
    try:
        with h5py.File(CONFIG['h5_filename'], 'r') as f:
            group_name = f"group_{code[:2]}"
            if group_name in f and code in f[group_name] and yesterday in f[group_name][code]:
                indicators_data = f[group_name][code][yesterday]['indicators'][()]
                if isinstance(indicators_data, bytes):
                    indicators = json.loads(indicators_data.decode('utf-8'))
                else:
                    indicators = json.loads(indicators_data)
                return indicators.get('price', 0.0)
    except:
        pass

    return 0.0

# ==================== ✅ 新增：涨跌停判断增强函数 ====================

def check_limit_status(code, price, pre_close, a_vol, b_vol, z_bias):
    """
    判断涨跌停状态（增强版）
    返回: (is_limit_down, is_limit_up, z_bias_fixed, imbalance_fixed)
    """
    is_limit_down = False
    is_limit_up = False
    z_bias_fixed = z_bias
    imbalance_fixed = None  # None表示不修改

    if pre_close <= 0 or price <= 0:
        return is_limit_down, is_limit_up, z_bias_fixed, imbalance_fixed

    # 获取股票名称用于ST判断
    name = get_stock_name(code)

    # 计算精确的涨跌停价（考虑不同板块）
    if str(code).startswith(('300', '301', '688')):  # 科创板/创业板 20%
        limit_down_price = round(pre_close * 0.8, 2)
        limit_up_price = round(pre_close * 1.2, 2)
    elif 'ST' in name or '*ST' in name:  # ST股票 5%
        limit_down_price = round(pre_close * 0.95, 2)
        limit_up_price = round(pre_close * 1.05, 2)
    else:  # 主板 10%
        limit_down_price = round(pre_close * 0.9, 2)
        limit_up_price = round(pre_close * 1.1, 2)

    # ✅ 跌停判断（严格）
    if abs(price - limit_down_price) <= 0.01:  # 价格等于跌停价（允许1分误差）
        if a_vol == 0 and b_vol > 0:  # 卖一为0，买一有挂单
            is_limit_down = True
            z_bias_fixed = -1.0
            imbalance_fixed = -1.0

    # ✅ 涨停判断（严格）- 使用 elif 确保互斥
    elif abs(price - limit_up_price) <= 0.01:  # 价格等于涨停价（允许1分误差）
        if b_vol == 0 and a_vol > 0:  # 买一为0，卖一有挂单
            is_limit_up = True
            z_bias_fixed = 1.0
            imbalance_fixed = 1.0

    return is_limit_down, is_limit_up, z_bias_fixed, imbalance_fixed

# 初始化股票名称
print(f"{Fore.CYAN}🔄 正在加载股票名称库...")
load_stock_names_from_tdx()

# ==================== 核心功能函数 ====================

def get_best_server():
    """获取最佳服务器"""
    servers = [
        {'ip': '119.147.212.81', 'port': 7709, 'name': '深圳主力'},
        {'ip': '114.80.149.19', 'port': 7709, 'name': '上海双线'},
        {'ip': '218.75.126.9', 'port': 7709, 'name': '长沙电信'},
    ]

    for retry in range(2):
        for s in servers:
            try:
                api = TdxHq_API()
                if api.connect(s['ip'], s['port'], time_out=2):
                    return api
            except:
                continue
        time.sleep(0.5)

    return None

def get_smoothed_imbalance(api, market, code):
    """获取平滑委比（五档）"""
    raw_imbalance = 0.0
    b_vol = a_vol = 0

    try:
        quotes = api.get_security_quotes([(market, code)])
        if quotes and len(quotes) > 0:
            quote = quotes[0]
            bid1 = int(quote.get('bid_vol1', 0))
            bid2 = int(quote.get('bid_vol2', 0))
            bid3 = int(quote.get('bid_vol3', 0))
            bid4 = int(quote.get('bid_vol4', 0))
            bid5 = int(quote.get('bid_vol5', 0))
            ask1 = int(quote.get('ask_vol1', 0))
            ask2 = int(quote.get('ask_vol2', 0))
            ask3 = int(quote.get('ask_vol3', 0))
            ask4 = int(quote.get('ask_vol4', 0))
            ask5 = int(quote.get('ask_vol5', 0))

            b_vol = bid1 + bid2 + bid3 + bid4 + bid5
            a_vol = ask1 + ask2 + ask3 + ask4 + ask5
            denominator = b_vol + a_vol

            if denominator > 0:
                raw_imbalance = (b_vol - a_vol) / denominator
    except Exception as e:
        if CONFIG['debug_mode']:
            print(f"{Fore.YELLOW}⚠️ 获取委比失败 {code}: {e}")

    with imbalance_history_lock:
        if code not in imbalance_history:
            imbalance_history[code] = deque(maxlen=CONFIG['imbalance_smooth_window'])
        imbalance_history[code].append(raw_imbalance)
        values = list(imbalance_history[code])
        smooth = sum(values) / len(values) if values else 0.0

    return smooth, raw_imbalance, b_vol, a_vol

# ==================== 量额比计算（修正版）====================

def calculate_volume_price_ratio(df, current_price):
    """
    计算量额比 = 当前价 / 成交均价
    含义：当前成交价格相对于全天平均成本的溢价/折价率
    >1 : 溢价买入，主力拉升
    =1 : 平价成交
    <1 : 折价卖出，主力压价
    """
    if df.empty or current_price <= 0:
        return 1.0, 1.0, 1.0, '⚪ 未知'

    try:
        df = df.copy()
        df['成交额'] = df['价格'] * df['成交量'] * 100

        # 计算每股成交均价
        mask = df['成交量'] > 0
        df.loc[mask, '成交均价'] = df.loc[mask, '成交额'] / (df.loc[mask, '成交量'] * 100)
        df.loc[~mask, '成交均价'] = df.loc[~mask, '价格']

        # 修正：当前价 / 成交均价
        df['量额比'] = current_price / df['成交均价']

        # 加权平均量额比（按成交额加权）
        total_amount = df['成交额'].sum()
        if total_amount > 0:
            weighted_vpr = (df['量额比'] * df['成交额']).sum() / total_amount
        else:
            weighted_vpr = 1.0

        # 分别计算买盘和卖盘的量额比
        if '买卖' in df.columns:
            buy_df = df[df['买卖'] == '买盘']
            sell_df = df[df['买卖'] == '卖盘']

            buy_amount = buy_df['成交额'].sum() if not buy_df.empty else 0
            sell_amount = sell_df['成交额'].sum() if not sell_df.empty else 0

            buy_vpr = (buy_df['量额比'] * buy_df['成交额']).sum() / buy_amount if buy_amount > 0 else 1.0
            sell_vpr = (sell_df['量额比'] * sell_df['成交额']).sum() / sell_amount if sell_amount > 0 else 1.0
        else:
            buy_vpr = sell_vpr = weighted_vpr

        # 主力信心评级
        if weighted_vpr > 1.015:
            confidence = '🔥 强拉升'
        elif weighted_vpr > 1.005:
            confidence = '📈 温和拉升'
        elif weighted_vpr < 0.985:
            confidence = '💧 强压价'
        elif weighted_vpr < 0.995:
            confidence = '📉 温和压价'
        else:
            confidence = '⚪ 均衡'

        return round(weighted_vpr, 4), round(buy_vpr, 4), round(sell_vpr, 4), confidence

    except Exception as e:
        if CONFIG['debug_mode']:
            print(f"{Fore.YELLOW}⚠️ 量额比计算失败: {e}")
        return 1.0, 1.0, 1.0, '⚪ 未知'

def update_score_history(code, score):
    """更新分数历史"""
    with score_history_lock:
        if code not in score_history:
            score_history[code] = deque(maxlen=3)
        score_history[code].append(score)

def get_score_trend_symbol(code):
    """获取分数趋势符号"""
    with score_history_lock:
        if code not in score_history:
            return " "
        scores = list(score_history[code])
        if len(scores) < 2:
            return " "
        last = scores[-1]
        prev = scores[-2]
        if last > prev + 3:
            return "↑"
        elif last < prev - 3:
            return "↓"
        else:
            return "→"

def calculate_stock_score(stock_data, intent, z_bias, imbalance=0.0, vpr=1.0):
    """计算股票信号分数"""
    score = 30

    # 趋势分数
    trend = stock_data.get('trend', '')
    if trend == "📈上移":
        score += 20
    elif trend == "🧱稳固":
        score += 10

    # 意图分数
    if "多头滚动买入" in intent:
        score += 35
    elif "机构吸筹" in intent:
        score += 30
    elif "短期滚动强势" in intent:
        score += 25
    elif "震仓洗盘" in intent:
        score += 15
    elif "均衡" in intent:
        score += 15
    elif "弱势震荡" in intent:
        score += 5
    elif "空头滚动卖出" in intent:
        score -= 15
    elif "机构减仓" in intent:
        score -= 25
    elif "筹码洗刷" in intent:
        score -= 10

    # 博弈比分数
    score += calculate_z_score(z_bias)

    # 下午时段补偿
    if CONFIG['enable_z_bias_time_compensation']:
        now = datetime.datetime.now()
        if now.hour >= 13:
            z_score = calculate_z_score(z_bias)
            compensated_score = int(z_score * CONFIG['z_bias_afternoon_multiplier'])
            score += (compensated_score - z_score)

    # 委比加减分
    if CONFIG['enable_order_imbalance']:
        if imbalance > CONFIG['imbalance_strong_buy']:
            score += CONFIG['imbalance_buy_score']
        elif imbalance < CONFIG['imbalance_strong_sell']:
            score -= CONFIG['imbalance_sell_score']

    # 量额比加减分
    if CONFIG.get('enable_volume_price_ratio', True):
        if vpr > CONFIG.get('vpr_strong_buy', 1.01):
            bonus = int((vpr - 1.01) * 1000)
            score += min(CONFIG.get('vpr_buy_score', 15), max(0, bonus))
        elif vpr < CONFIG.get('vpr_strong_sell', 0.99):
            penalty = int((0.99 - vpr) * 1000)
            score -= min(CONFIG.get('vpr_sell_score', 15), max(0, penalty))

    return max(0, min(100, score))

def analyze_intent(df_today, z_bias):
    """分析主力意图"""
    if df_today is None or df_today.empty:
        return "⚖️ 均衡", 0

    if z_bias >= 0.20:
        intent = "📈 多头滚动买入"
    elif z_bias >= 0.12:
        intent = "🕵️ 机构吸筹"
    elif z_bias >= 0.05:
        intent = "📊 短期滚动强势"
    elif z_bias >= 0.02:
        intent = "🧱 震仓洗盘"
    elif z_bias <= -0.20:
        intent = "⚠️ 机构减仓"
    elif z_bias <= -0.12:
        intent = "📉 空头滚动卖出"
    elif z_bias <= -0.05:
        intent = "🌪️ 筹码洗刷"
    elif z_bias <= -0.02:
        intent = "🧱 弱势震荡"
    else:
        intent = "⚖️ 均衡"

    return intent, 0

def h5_get_peak_by_date(code, date):
    """获取指定日期峰值"""
    if not os.path.exists(CONFIG['h5_filename']):
        return None

    try:
        with h5py.File(CONFIG['h5_filename'], 'r') as f:
            group_name = f"group_{code[:2]}"
            if group_name not in f or code not in f[group_name] or date not in f[group_name][code]:
                return None

            indicators_data = f[group_name][code][date]['indicators'][()]
            if isinstance(indicators_data, bytes):
                indicators_str = indicators_data.decode('utf-8')
            else:
                indicators_str = str(indicators_data)
            indicators = json.loads(indicators_str)
            return float(indicators.get('peak', 0))
    except:
        return None

def h5_get_previous_peak(code, today):
    """优先查今日已有峰值，再查昨日"""
    today_peak = h5_get_peak_by_date(code, today)
    if today_peak is not None and today_peak > 0:
        return today_peak

    yesterday = (datetime.date.today() - datetime.timedelta(days=1)).strftime('%Y-%m-%d')
    yesterday_peak = h5_get_peak_by_date(code, yesterday)
    if yesterday_peak is not None:
        return yesterday_peak

    for i in range(2, 5):
        d = (datetime.date.today() - datetime.timedelta(days=i)).strftime('%Y-%m-%d')
        peak = h5_get_peak_by_date(code, d)
        if peak is not None:
            return peak

    return None

def h5_load_transactions(code, date):
    """加载历史交易数据"""
    if not os.path.exists(CONFIG['h5_filename']):
        return pd.DataFrame()

    try:
        with h5py.File(CONFIG['h5_filename'], 'r') as f:
            group_name = f"group_{code[:2]}"
            if group_name not in f:
                return pd.DataFrame()
            if code not in f[group_name]:
                return pd.DataFrame()
            if date not in f[group_name][code]:
                return pd.DataFrame()

            data = f[group_name][code][date]['transactions'][:]

            df = pd.DataFrame({
                'time': [x.decode('utf-8') if isinstance(x, bytes) else str(x) for x in data['time']],
                '价格': [float(x) for x in data['价格']],
                '成交量': [float(x) for x in data['成交量']],
                'bs': [int(x) for x in data['bs']],
                '买卖': [x.decode('utf-8') if isinstance(x, bytes) else str(x) for x in data['买卖']],
            })
            return df
    except:
        return pd.DataFrame()

def h5_save_data(code, date, df, indicators):
    """保存数据到HDF5"""
    try:
        with h5_write_lock:
            with h5py.File(CONFIG['h5_filename'], 'a', libver='latest') as f:
                group_name = f"group_{code[:2]}"
                if group_name not in f:
                    f.create_group(group_name)
                if code not in f[group_name]:
                    f[group_name].create_group(code)
                if date in f[group_name][code]:
                    del f[group_name][code][date]

                date_group = f[group_name][code].create_group(date)

                if not df.empty:
                    df_save = df.tail(200)
                    dtype = np.dtype([
                        ('time', h5py.string_dtype()),
                        ('价格', 'f8'),
                        ('成交量', 'f8'),
                        ('bs', 'i4'),
                        ('买卖', h5py.string_dtype()),
                    ])
                    data_array = np.array([
                        (str(row['time']), float(row['价格']), float(row['成交量']),
                         int(row['bs']), str(row['买卖']))
                        for _, row in df_save.iterrows()
                    ], dtype=dtype)
                    date_group.create_dataset('transactions', data=data_array, compression='gzip')

                indicators_json = json.dumps(indicators, ensure_ascii=False)
                dt = h5py.special_dtype(vlen=str)
                date_group.create_dataset('indicators',
                                          data=np.array(indicators_json, dtype=object),
                                          dtype=dt)

        with performance_lock:
            performance_stats['h5_write_success'] += 1
        return True
    except:
        with performance_lock:
            performance_stats['h5_write_failed'] += 1
        return False

# ==================== ✅ 修复：共振连续天数计算函数 ====================

def calculate_resonance_streak(code, dates_list):
    """
    根据历史共振日期计算真正的连续天数
    参数:
        code: 股票代码
        dates_list: 按时间顺序排列的日期列表（从早到晚）
    返回:
        连续天数
    """
    if not dates_list:
        return 0

    # 从最新的日期开始往前推
    streak = 1
    for i in range(len(dates_list) - 1, 0, -1):
        current_date = datetime.datetime.strptime(dates_list[i], '%Y-%m-%d')
        prev_date = datetime.datetime.strptime(dates_list[i - 1], '%Y-%m-%d')

        # 计算日期差（容忍周末）
        days_diff = (current_date - prev_date).days

        if days_diff <= 3:  # 连续（包括周末）
            streak += 1
        else:
            break

        # 最多计算5天
        if streak >= RESONANCE_CONFIG['streak_days']:
            break

    return min(streak, RESONANCE_CONFIG['streak_days'])

# ==================== 共振连续历史管理（修复版）====================

def get_streak_bar(days, max_days=5):
    """获取连续天数热力条"""
    if days == 0:
        return "░░░░░"
    filled = "█" * days
    empty = "░" * (max_days - days)
    if days >= 5:
        color = Fore.RED
    elif days >= 3:
        color = Fore.YELLOW
    else:
        color = Fore.WHITE
    return f"{color}{filled}{empty}{Fore.RESET}"

def get_resonance_frequency(code, days=5):
    """获取最近N天的共振频率"""
    with resonance_history_lock:
        if code not in resonance_history:
            return 0, days
        dates = list(resonance_history[code])
        if not dates:
            return 0, days
        today = datetime.date.today()
        count = 0
        for i in range(days):
            date = (today - datetime.timedelta(days=i)).strftime('%Y-%m-%d')
            if date in dates:
                count += 1
        return count, days

def update_resonance_history(code, today, is_resonance):
    """
    更新核心共振连续上榜记录（修复版）
    功能：连续记录、基于实际日期计算连续天数、上限保护、连续报警
    """
    with resonance_history_lock:
        # 初始化
        if code not in resonance_history:
            resonance_history[code] = deque(maxlen=RESONANCE_CONFIG['streak_days'])

        if is_resonance:
            # 本轮上榜：记录日期
            if len(resonance_history[code]) == 0:
                # 首次上榜
                resonance_history[code].append(today)
                resonance_streak[code] = 1
            else:
                # 检查是否与最新记录连续
                last_date = resonance_history[code][-1]
                last_dt = datetime.datetime.strptime(last_date, '%Y-%m-%d')
                today_dt = datetime.datetime.strptime(today, '%Y-%m-%d')

                # 判断是否连续交易日（容忍周末）
                if (today_dt - last_dt).days <= 3:
                    resonance_history[code].append(today)
                    # ✅ 修复：重新计算真正的连续天数
                    resonance_streak[code] = calculate_resonance_streak(code, list(resonance_history[code]))
                else:
                    # 中断，重新开始
                    resonance_history[code].clear()
                    resonance_history[code].append(today)
                    resonance_streak[code] = 1

            # 连续天数报警
            name = get_stock_name(code)
            if resonance_streak[code] == 3:
                print(f"{Fore.RED}🔥🔥 连续3天共振！强庄确认：{code} {name} 🔥🔥{Fore.RESET}")
            elif resonance_streak[code] == 5:
                print(f"{Fore.RED}👑👑👑 连续5天共振！绝对龙头：{code} {name} 👑👑👑{Fore.RESET}")
        else:
            # 本轮未上榜：如果超过3天未上榜，清零历史
            if code in resonance_history and len(resonance_history[code]) > 0:
                last_date = resonance_history[code][-1]
                last_dt = datetime.datetime.strptime(last_date, '%Y-%m-%d')
                today_dt = datetime.datetime.strptime(today, '%Y-%m-%d')

                if (today_dt - last_dt).days > 3:
                    resonance_history[code].clear()
                    resonance_streak[code] = 0

def get_resonance_streak_display(code):
    """获取共振连续天数显示格式"""
    with resonance_history_lock:
        days = resonance_streak.get(code, 0)
        if days == 0:
            return "  ", Fore.WHITE

        symbols = {1: "Ⅰ", 2: "Ⅱ", 3: "Ⅲ", 4: "Ⅳ", 5: "Ⅴ"}
        symbol = symbols.get(days, f"{days}天")
        color = RESONANCE_CONFIG['streak_colors'].get(days, Fore.WHITE)

        return f"{symbol}", color

def get_top_resonance_streaks(limit=10):
    """获取连续上榜天数排行榜"""
    with resonance_history_lock:
        streaks = [(code, days) for code, days in resonance_streak.items() if days > 0]
        # 按连续天数降序，同天数按最新上榜时间
        streaks.sort(key=lambda x: (-x[1],
                                    resonance_history[x[0]][-1] if resonance_history[x[0]] else ''),
                     reverse=False)
        return streaks[:limit]

# ==================== 多信号组合矩阵策略 ====================

def check_signal_combinations(code, stock_data, result, top_signals_list, scan_count):
    """检查多信号组合"""
    triggered_combinations = []
    today = datetime.date.today().strftime('%Y-%m-%d')

    is_resonance = result.get('is_resonance', False)
    is_strong = result.get('is_strong', False)
    is_warning = result.get('is_warning', False)
    vpr = stock_data.get('vpr', 1.0)
    z_bias = stock_data.get('z_bias', 0)

    is_top1 = False
    if top_signals_list and len(top_signals_list) > 0:
        is_top1 = (top_signals_list[0].get('code') == code)

    is_new_top10 = False
    if top_signals_list:
        codes_in_top10 = [s.get('code') for s in top_signals_list[:10]]
        is_new_top10 = code in codes_in_top10

    vpr_strong_buy = vpr > CONFIG.get('vpr_strong_buy', 1.01)
    vpr_strong_sell = vpr < CONFIG.get('vpr_strong_sell', 0.99)
    vpr_mild_buy = vpr > CONFIG.get('vpr_mild_buy', 1.005)
    z_bias_severe_sell = z_bias < CONFIG.get('z_bias_severe_sell', -0.2)

    streak_days = resonance_streak.get(code, 0)
    streak_3 = streak_days >= 3
    streak_5 = streak_days >= 5

    conditions = {
        'is_resonance': is_resonance,
        'is_strong': is_strong,
        'is_warning': is_warning,
        'is_top1': is_top1,
        'is_new_top10': is_new_top10,
        'vpr_strong_buy': vpr_strong_buy,
        'vpr_strong_sell': vpr_strong_sell,
        'vpr_mild_buy': vpr_mild_buy,
        'z_bias_severe_sell': z_bias_severe_sell,
        'streak_3': streak_3,
        'streak_5': streak_5
    }

    with combination_history_lock:
        for combo_key, combo_config in SIGNAL_COMBINATION_CONFIG['combinations'].items():
            all_conditions_met = all(conditions.get(cond, False) for cond in combo_config['conditions'])

            if all_conditions_met:
                combo_id = f"{code}_{combo_key}_{today}_{scan_count}"

                if combo_id not in combination_history:
                    combination_history[combo_id] = {
                        'time': datetime.datetime.now().isoformat(),
                        'code': code,
                        'name': stock_data.get('name', ''),
                        'combination': combo_key,
                        'strategy': combo_config['strategy'],
                        'confidence': combo_config['confidence']
                    }

                    triggered_combinations.append({
                        'code': code,
                        'name': stock_data.get('name', ''),
                        'combo_key': combo_key,
                        'combo_name': combo_config['name'],
                        'confidence': combo_config['confidence'],
                        'strategy': combo_config['strategy'],
                        'color': combo_config['color'],
                        'alert': combo_config.get('alert', False)
                    })

    return triggered_combinations

# ==================== 显示管理类（优化版）====================

class DisplayManager:
    """显示管理类 - 优化对齐，解决字符拥挤"""

    @staticmethod
    def print_header():
        width = DISPLAY_CONFIG['terminal_width']
        mode = "紧凑" if DISPLAY_CONFIG['compact_mode'] else "标准"
        print(f"\n{Fore.CYAN}{'主力监测雷达 V5.8.5 (共振连续榜计算修复版)':^{width}}")
        print(f"{Fore.CYAN}{f'量额比(VPR) | 拉升溢价 | 压价折价 | 完整续跑 | 涨跌停精准(增强) | 共振连续5日榜(修复) | {mode}模式':^{width}}")
        print(f"{Fore.CYAN}{'=' * width}")
        print(f"{Fore.GREEN}✅ 量额比修正：当前价/成交均价 | 红色=溢价拉升 | 绿色=折价压价{Fore.RESET}")
        print(f"{Fore.GREEN}✅ 共振榜修复：显示当前共振意图，而非历史意图{Fore.RESET}")
        print(f"{Fore.GREEN}✅ 涨跌停增强：昨收盘价4层获取 | 板块差异识别 | 互斥判断{Fore.RESET}")
        print(f"{Fore.GREEN}✅ 连续天数修复：基于实际连续日期计算，频率与连续天数一致{Fore.RESET}")

    @staticmethod
    def print_scan_info(scan_count, time_str):
        width = DISPLAY_CONFIG['terminal_width']
        if scan_count == 1 or scan_count % 10 == 0:
            if SESSION_STATE.get('scan_count', 0) > 0 and scan_count == 1:
                prefix = "🔄 续跑"
            else:
                prefix = f"第{scan_count}轮"
            print(f"\n{Fore.YELLOW}{f'{prefix}扫描 [{time_str}]':^{width}}")

    @staticmethod
    def print_resonance_header():
        width = DISPLAY_CONFIG['terminal_width']
        print(f"\n{Fore.RED}{'🎯 核心共振信号 🎯':^{width}}")
        print(f"{Fore.RED}{'-' * width}")

    @staticmethod
    def print_warning_header():
        width = DISPLAY_CONFIG['terminal_width']
        print(f"\n{Fore.MAGENTA}{'🚨 风险报警信号 🚨':^{width}}")
        print(f"{Fore.MAGENTA}{'-' * width}")

    @staticmethod
    def print_strong_header():
        width = DISPLAY_CONFIG['terminal_width']
        print(f"\n{Fore.YELLOW}{'📢 强势买入信号 📢':^{width}}")
        print(f"{Fore.YELLOW}{'-' * width}")

    @staticmethod
    def print_top_signals_header():
        width = DISPLAY_CONFIG['terminal_width']
        w = DISPLAY_CONFIG['width']
        sp = ' ' * DISPLAY_CONFIG['spacing']

        header_parts = [
            f"{'排名':<{w['rank']}}",
            f"{'代码':<{w['code']}}",
            f"{'名称':<{w['name']}}",
            f"{'趋':<{w['trend']}}",
            f"{'意图':<{w['intent']}}",
            f"{'计':<{w['count']}}",
            f"{'分数':<{w['score']}}",
            f"{'趋':<{w['trend']}}",
            f"{'Z偏':<{w['z_bias']}}",
            f"{'委比':<{w['imbalance']}}",
            f"{'量额':<{w['vpr']}}",
            f"{'价格':<{w['price']}}",
        ]

        if w['streak'] > 0:
            header_parts.append(f"{'连':<{w['streak']}}")
        if w['bar'] > 0:
            header_parts.append(f"{'热力':<{w['bar']}}")

        header = sp.join(header_parts)

        print(f"\n{Fore.CYAN}{'🏆 顶部信号排行 🏆':^{width}}")
        print(f"{Fore.CYAN}{'-' * width}")
        print(f"{Fore.CYAN}{header}")
        print(f"{Fore.CYAN}{'-' * width}")

    @staticmethod
    def print_stock_info(stock_data, signal_type='normal', rank=None):
        if stock_data is None:
            return

        code = stock_data.get('code', '')
        name = get_stock_name(code)
        intent = stock_data.get('intent', '')
        trend = stock_data.get('trend', '')
        z_bias = stock_data.get('z_bias', 0)
        imbalance = stock_data.get('imbalance', 0.0)
        price = stock_data.get('price', 0)
        count = stock_data.get('count', 0)
        score = stock_data.get('total_score', 30)
        vpr = stock_data.get('vpr', 1.0)

        signal_config = {
            'resonance': {'symbol': '🎯', 'color': Fore.RED + Style.BRIGHT},
            'warning': {'symbol': '🚨', 'color': Fore.MAGENTA + Style.BRIGHT},
            'strong': {'symbol': '📢', 'color': Fore.YELLOW + Style.BRIGHT},
            'top': {'symbol': '🏆', 'color': Fore.CYAN},
            'normal': {'symbol': '⚖', 'color': Fore.WHITE},
        }
        cfg = signal_config.get(signal_type, signal_config['normal'])
        symbol = cfg['symbol']
        color = cfg['color']

        intent_map = {
            "📈 多头滚动买入": "多头",
            "🕵️ 机构吸筹": "吸筹",
            "📊 短期滚动强势": "短强",
            "🧱 震仓洗盘": "震仓",
            "⚠️ 机构减仓": "减仓",
            "📉 空头滚动卖出": "空头",
            "🌪️ 筹码洗刷": "洗刷",
            "🧱 弱势震荡": "弱势",
            "⚖️ 均衡": "均衡",
            "🔒 跌停": "跌停",
            "📈 涨停": "涨停",
            "🔒 跌停(估)": "跌停",
            "📈 涨停(估)": "涨停",
        }
        intent_short = intent_map.get(intent, intent[:2]) if intent else "  "
        if len(intent_short) > 4:
            intent_short = intent_short[:4]

        trend_symbol = {'📈上移': '↗', '📉下移': '↘', '🧱稳固': '●'}.get(trend, ' ')

        streak_days = resonance_streak.get(code, 0)
        king_mark = '👑' * (streak_days // 2) if streak_days >= 3 else ''

        name_display = trim_visual(name + king_mark, DISPLAY_CONFIG['width']['name'])
        name_display = pad_visual(name_display, DISPLAY_CONFIG['width']['name'])

        w = DISPLAY_CONFIG['width']
        sp = ' ' * DISPLAY_CONFIG['spacing']

        code_display = f"{code:<{w['code']}}"
        trend_display = f"{trend_symbol:<{w['trend']}}"
        intent_display = f"{intent_short:<{w['intent']}}"
        count_display = f"{count:>{w['count']}}" if count > 0 else ' ' * w['count']
        score_display = f"S{score:03d}"[-w['score']:].rjust(w['score'])

        trend_symbol_score = get_score_trend_symbol(code)
        trend_score_display = f"{trend_symbol_score:<{w['trend']}}"

        z_display = f"Z{z_bias:+.3f}"[:w['z_bias']].rjust(w['z_bias'])

        if abs(imbalance) > 0.001:
            imb_display = f"{imbalance:+.3f}"[:w['imbalance']].rjust(w['imbalance'])
        else:
            imb_display = '0.000'.rjust(w['imbalance'])

        if abs(vpr - 1.0) > 0.001:
            vpr_display = f"{vpr:.3f}"[:w['vpr']].rjust(w['vpr'])
        else:
            vpr_display = '1.000'.rjust(w['vpr'])

        price_display = f"{price:>7.2f}"[:w['price']].rjust(w['price'])

        # 量额比颜色规则
        if vpr > 1.01:
            vpr_color = Fore.RED
        elif vpr < 0.99:
            vpr_color = Fore.GREEN
        else:
            vpr_color = Fore.WHITE

        parts = []

        if signal_type == 'top':
            rank_display = f"{rank:2d}." if rank else '   '
            parts.append(f"{color}{rank_display}")

        parts.extend([
            f"{code_display}",
            f"{name_display}",
            f"{trend_display}",
            f"{intent_display}",
            f"{count_display}",
            f"{score_display}",
            f"{trend_score_display}",
            f"{z_display}",
            f"{imb_display}",
            f"{vpr_color}{vpr_display}{Fore.RESET}",
            f"{price_display}",
        ])

        if w['streak'] > 0:
            streak_symbol, streak_color = get_resonance_streak_display(code)
            streak_display = f"{streak_color}{streak_symbol}{Fore.RESET}"
            streak_display = pad_visual(streak_display, w['streak'])
            parts.append(f"{streak_display}")

        if w['bar'] > 0:
            bar = get_streak_bar(streak_days)
            bar_display = pad_visual(bar, w['bar'])
            parts.append(f"{bar_display}")

        line = sp.join(parts)
        print(f"{color}{line}{Fore.RESET}")

    @staticmethod
    def print_resonance_streak_board():
        """打印共振连续榜（修复：意图显示对齐，连续天数计算正确）"""
        streaks = get_top_resonance_streaks(10)
        if not streaks:
            return

        width = DISPLAY_CONFIG['terminal_width']
        w = DISPLAY_CONFIG['width']
        sp = ' ' * DISPLAY_CONFIG['spacing']

        header_parts = [
            f"{'排名':<4}",
            f"{'代码':<{w['code']}}",
            f"{'名称':<{w['name']}}",
            f"{'连续':<{w['streak']+2}}",
            f"{'强度':<7}",
            f"{'频率':<6}",
            f"{'最新价':<{w['price']}}",
            f"{'最新意图':<{10}}",  # 固定视觉宽度10
        ]
        if w['bar'] > 0:
            header_parts.append(f"{'热力条':<{w['bar']}}")

        header = sp.join(header_parts)

        print(f"\n{Fore.RED}{'🔥 共振连续榜（连续5日为强庄） 🔥':^{width}}")
        print(f"{Fore.RED}{'-' * width}")
        print(f"{Fore.RED}{header}")
        print(f"{Fore.RED}{'-' * width}")

        for i, (code, days) in enumerate(streaks, 1):
            name = get_stock_name(code)
            name_display = trim_visual(name, w['name'])
            name_display = pad_visual(name_display, w['name'])

            latest_price = get_latest_price(code)
            price_display = f"{latest_price:>7.2f}"[:w['price']].rjust(w['price'])

            # 获取当前共振意图
            latest_intent = get_latest_resonance_intent(code)
            latest_intent = latest_intent.replace('📈', '').replace('🕵️', '').replace('📊', '')
            latest_intent = latest_intent.replace('🧱', '').replace('⚠️', '').replace('📉', '')
            latest_intent = latest_intent.replace('🌪️', '').replace('🔒', '').replace('📈', '')
            latest_intent = latest_intent.strip()

            # 按视觉宽度截取，最大10
            latest_intent = trim_visual(latest_intent, 10)
            if not latest_intent:
                latest_intent = '⚖️均衡'

            # 填充到固定视觉宽度
            latest_intent_display = pad_visual(latest_intent, 10)

            symbol, color = get_resonance_streak_display(code)
            strength = "🔴" * min(days, 3) + "⚪" * (3 - min(days, 3))

            freq_count, freq_days = get_resonance_frequency(code)
            frequency = f"{freq_count}/{freq_days}"

            parts = [
                f"{color}{i:2d}.",
                f"{code:<{w['code']}}",
                f"{name_display}",
                f"{symbol} {days}天",
                f"{strength:<7}",
                f"{frequency:<6}",
                f"{price_display}",
                f"{latest_intent_display}",  # 使用填充后的字符串
            ]

            if w['bar'] > 0:
                bar = get_streak_bar(days)
                bar_display = pad_visual(bar, w['bar'])
                parts.append(f"{bar_display}")

            line = sp.join(parts)
            print(f"{color}{line}{Fore.RESET}")

    @staticmethod
    def print_combination_signals(combinations, max_display=10):
        if not combinations:
            return

        width = DISPLAY_CONFIG['terminal_width']
        w = DISPLAY_CONFIG['width']
        sp = ' ' * DISPLAY_CONFIG['spacing']

        header_parts = [
            f"{'代码':<{w['code']}}",
            f"{'名称':<{w['name']}}",
            f"{'组合信号':<{w['combo_name']}}",
            f"{'置信度':<{w['confidence']}}",
            f"{'策略':<{w['strategy']}}",
            f"{'价格':<{w['price']}}",
            f"{'VPR':<{w['vpr']}}",
            f"{'Z偏':<{w['z_bias']}}",
        ]
        header = sp.join(header_parts)

        print(f"\n{Fore.GREEN}{'🎯 多信号组合策略信号 🎯':^{width}}")
        print(f"{Fore.GREEN}{'-' * width}")
        print(f"{Fore.GREEN}{header}")
        print(f"{Fore.GREEN}{'-' * width}")

        for combo in combinations[:max_display]:
            code = combo['code']
            name = get_stock_name(code)
            name_display = trim_visual(name, w['name'])
            name_display = pad_visual(name_display, w['name'])

            combo_name = trim_visual(combo['combo_name'], w['combo_name'])
            combo_name = pad_visual(combo_name, w['combo_name'])

            confidence = pad_visual(combo['confidence'], w['confidence'])
            strategy = pad_visual(combo['strategy'], w['strategy'])

            price = get_latest_price(code)
            price_display = f"{price:>7.2f}"[:w['price']].rjust(w['price'])

            with stock_data_cache_lock:
                if code in stock_data_cache:
                    vpr = stock_data_cache[code].get('vpr', 1.0)
                    z_bias = stock_data_cache[code].get('z_bias', 0.0)
                else:
                    vpr, z_bias = 1.0, 0.0

            vpr_display = f"{vpr:.3f}"[:w['vpr']].rjust(w['vpr'])
            z_display = f"{z_bias:+.3f}"[:w['z_bias']].rjust(w['z_bias'])

            vpr_color = Fore.RED if vpr > 1.01 else Fore.GREEN if vpr < 0.99 else Fore.WHITE
            color = combo['color']

            parts = [
                f"{color}{code:<{w['code']}}",
                f"{name_display}",
                f"{combo_name}",
                f"{confidence}",
                f"{strategy}",
                f"{price_display}",
                f"{vpr_color}{vpr_display}{Fore.RESET}",
                f"{z_display}",
            ]
            line = sp.join(parts)
            print(f"{color}{line}{Fore.RESET}")

    @staticmethod
    def print_stats(stats, warning_stats, elapsed_time=None, scan_count=0):
        width = DISPLAY_CONFIG['terminal_width']

        print(f"\n{Fore.CYAN}{'📊 本轮统计':^{width}}")

        stats_line = f"{Fore.RED}核心共振: {stats['resonance']:2d}  "
        stats_line += f"{Fore.MAGENTA}风险报警: {stats['warning']:2d}  "
        stats_line += f"{Fore.YELLOW}强势信号: {stats['strong']:2d}  "
        stats_line += f"{Fore.GREEN}正面信号: {stats['positive']:2d}  "

        with resonance_history_lock:
            streak_count = len([d for d in resonance_streak.values() if d > 0])
            streak_3plus = len([d for d in resonance_streak.values() if d >= 3])

        with combination_history_lock:
            today = datetime.date.today().strftime('%Y-%m-%d')
            combo_count = len([v for v in combination_history.values()
                               if v.get('time', '').startswith(today)])

        stats_line += f"{Fore.RED}连续共振: {streak_count:2d}只(3天+:{streak_3plus})  "
        stats_line += f"{Fore.GREEN}组合信号: {combo_count:2d}个"
        print(stats_line)

        if elapsed_time:
            print(f"{Fore.CYAN}耗时: {elapsed_time:.1f}秒")

# ==================== 实时处理函数（增强版）====================

def process_stock(code, api, scan_count, top_signals_list=None):
    """处理单只股票（涨跌停判断增强版）"""
    try:
        code = str(code).zfill(6)
        today = datetime.date.today().strftime('%Y-%m-%d')
        market = 1 if code.startswith(('6', '688')) else 0

        alerted_today = set()
        try:
            with h5py.File(CONFIG['h5_filename'], 'r') as f:
                if 'alerted' in f and today in f['alerted']:
                    alerted_data = f['alerted'][today][()]
                    if isinstance(alerted_data, bytes):
                        alerted_today = set(json.loads(alerted_data.decode('utf-8')))
                    else:
                        alerted_today = set(json.loads(alerted_data))
        except:
            pass

        stock_name = get_stock_name(code)
        smooth_imbalance, raw_imbalance, b_vol, a_vol = get_smoothed_imbalance(api, market, code)

        df_full = pd.DataFrame()
        try:
            raw_data = api.get_transaction_data(market, code, 0, 200)
            if raw_data:
                df_new = pd.DataFrame(raw_data)
                if not df_new.empty:
                    df_new.rename(columns={'price': '价格', 'vol': '成交量', 'buyorsell': 'bs'}, inplace=True)
                    df_new['买卖'] = df_new['bs'].apply(lambda x: '买盘' if x == 0 else '卖盘')
                    df_new['time'] = df_new['time'].astype(str)

                    df_old = h5_load_transactions(code, today)
                    if not df_old.empty:
                        df_full = pd.concat([df_old, df_new]).drop_duplicates(
                            subset=['time', '价格'], keep='last'
                        ).tail(200)
                    else:
                        df_full = df_new.tail(200)
        except Exception as e:
            if CONFIG['debug_mode']:
                print(f"{Fore.YELLOW}⚠️ {code} 获取交易数据异常: {e}")

        price = 0.0
        peak = 0.0
        z_bias = 0.0
        vpr = 1.0
        buy_vpr = 1.0
        sell_vpr = 1.0
        confidence = '⚪ 未知'
        is_limit_down = False
        is_limit_up = False
        buy_vol = 0
        sell_vol = 0

        if not df_full.empty:
            try:
                price = float(df_full['价格'].iloc[-1])
                vpr, buy_vpr, sell_vpr, confidence = calculate_volume_price_ratio(df_full, price)

                buy_vol = df_full[df_full['买卖'] == '买盘']['成交量'].sum()
                sell_vol = df_full[df_full['买卖'] == '卖盘']['成交量'].sum()

                if buy_vol + sell_vol > 0:
                    z_bias = (buy_vol - sell_vol) / (buy_vol + sell_vol)
                else:
                    z_bias = 0.0
                z_bias = round(z_bias, 3)

                dist = df_full.groupby('价格')['成交量'].sum()
                if not dist.empty:
                    peak = float(dist.idxmax())

            except Exception as e:
                if CONFIG['debug_mode']:
                    print(f"{Fore.YELLOW}⚠️ {code} 数据计算异常: {e}")
                z_bias = 0.0
                vpr = 1.0
        else:
            z_bias = 0.0
            vpr = 1.0
            confidence = '⚪ 无成交'

        # ==================== 涨跌停判断（增强版）====================
        # 获取昨收盘价（多重保障）
        pre_close = get_pre_close(api, market, code)

        # 涨跌停判断（增强版）
        is_limit_down, is_limit_up, fixed_z_bias, fixed_imbalance = check_limit_status(
            code, price, pre_close, a_vol, b_vol, z_bias
        )

        # 应用修正值
        if is_limit_down or is_limit_up:
            z_bias = fixed_z_bias
            if fixed_imbalance is not None:
                smooth_imbalance = fixed_imbalance
                raw_imbalance = fixed_imbalance

        # 无昨收盘价时的降级判断
        if pre_close <= 0 and not is_limit_down and not is_limit_up:
            # 跌停特征：卖一为0，买一巨大，博弈比极负
            if a_vol == 0 and b_vol > 10000 and z_bias < -0.95:
                is_limit_down = True
                z_bias = -1.0
                smooth_imbalance = -1.0
                raw_imbalance = -1.0
            # 涨停特征：买一为0，卖一巨大，博弈比极正
            elif b_vol == 0 and a_vol > 10000 and z_bias > 0.95:
                is_limit_up = True
                z_bias = 1.0
                smooth_imbalance = 1.0
                raw_imbalance = 1.0

        # ==================== 趋势判断 ====================
        p_peak = h5_get_previous_peak(code, today)

        if p_peak and p_peak > 0 and peak > 0:
            shift_ratio = (peak - p_peak) / p_peak

            if price < 10:
                threshold = CONFIG['trend_threshold_low']
            elif price < 30:
                threshold = CONFIG['trend_threshold_medium']
            else:
                threshold = CONFIG['trend_threshold_high']

            if shift_ratio > threshold:
                trend = "📈上移"
            elif shift_ratio < -threshold:
                trend = "📉下移"
            else:
                trend = "🧱稳固"
        else:
            trend = "🧱稳固"

        # ==================== 意图分析 ====================
        if is_limit_down:
            if pre_close <= 0:
                intent = "🔒 跌停(估)"
            else:
                intent = "🔒 跌停"
            trend = "📉下移"
        elif is_limit_up:
            if pre_close <= 0:
                intent = "📈 涨停(估)"
            else:
                intent = "📈 涨停"
            trend = "📈上移"
        else:
            intent, _ = analyze_intent(df_full, z_bias)

        # ==================== 更新意图计数器 ====================
        with intent_counter_lock:
            if code not in intent_counter:
                intent_counter[code] = {}
            if intent not in intent_counter[code]:
                intent_counter[code][intent] = 0
            intent_counter[code][intent] += 1
            count = intent_counter[code][intent]

        # ==================== 准备股票数据 ====================
        stock_data = {
            'code': code,
            'name': stock_name,
            'intent': intent,
            'trend': trend,
            'z_bias': float(z_bias),
            'imbalance': smooth_imbalance,
            'price': price,
            'count': count,
            'vpr': vpr,
            'vpr_confidence': confidence,
        }

        # ==================== 计算分数 ====================
        if is_limit_down:
            score = 5
        elif is_limit_up:
            score = 80
        else:
            score = calculate_stock_score(stock_data, intent, z_bias, smooth_imbalance, vpr)

        stock_data['total_score'] = score
        update_score_history(code, score)

        # ==================== 信号判断 ====================
        if is_limit_down:
            is_resonance = False
            is_warning = True
            is_strong = False
            is_positive = False
            signal_level = 'warning'
        elif is_limit_up:
            is_resonance = False
            is_warning = False
            is_strong = True
            is_positive = False
            signal_level = 'strong'
        else:
            is_resonance = (
                    score >= CONFIG['resonance_score_threshold'] and
                    z_bias > CONFIG['resonance_z_bias_threshold'] and
                    intent in ["📈 多头滚动买入", "🕵️ 机构吸筹", "📊 短期滚动强势"]
            )

            is_warning = (
                    score <= CONFIG['warning_score_threshold'] or
                    intent in ["⚠️ 机构减仓", "📉 空头滚动卖出", "🌪️ 筹码洗刷"] or
                    z_bias < CONFIG['warning_z_bias_threshold']
            )

            is_strong = (
                    score >= CONFIG['strong_score_threshold'] and
                    z_bias > CONFIG['strong_z_bias_threshold'] and
                    intent in ["📈 多头滚动买入", "🕵️ 机构吸筹", "📊 短期滚动强势", "🧱 震仓洗盘"]
            )

            is_positive = (
                    score >= CONFIG['positive_score_threshold'] and
                    not is_warning and not is_resonance and not is_strong
            )

            if is_resonance:
                signal_level = 'resonance'
            elif is_warning:
                signal_level = 'warning'
            elif is_strong:
                signal_level = 'strong'
            elif is_positive:
                signal_level = 'positive'
            else:
                signal_level = 'normal'

        # ==================== 更新共振历史 ====================
        update_resonance_history(code, today, is_resonance)

        # ==================== 报警去重统计 ====================
        alert_key = f"{code}_{signal_level}"
        is_new_alert = alert_key not in alerted_today

        if is_warning and is_new_alert:
            with warning_stats_lock:
                warning_stats['total_warnings'] += 1
                warning_stats['warning_stocks'].add(code)
                if intent not in warning_stats['warning_by_type']:
                    warning_stats['warning_by_type'][intent] = 0
                warning_stats['warning_by_type'][intent] += 1

            try:
                alerted_today.add(alert_key)
                with h5py.File(CONFIG['h5_filename'], 'a') as f:
                    if 'alerted' not in f:
                        f.create_group('alerted')
                    dt = h5py.special_dtype(vlen=str)
                    if today in f['alerted']:
                        del f['alerted'][today]
                    f['alerted'].create_dataset(
                        today,
                        data=np.array(json.dumps(list(alerted_today)), dtype=object),
                        dtype=dt
                    )
            except Exception as e:
                if CONFIG['debug_mode']:
                    print(f"{Fore.YELLOW}⚠️ 保存报警记录失败: {e}")

        # ==================== 保存数据到H5和缓存 ====================
        indicators = {
            'code': code,
            'name': stock_name,
            'date': today,
            'price': price,
            'peak': peak,
            'trend': trend,
            'intent': intent,
            'z_bias': float(z_bias),
            'vpr': vpr,
            'score': int(score),
            'signal': signal_level,
            'is_limit_down': is_limit_down,
            'is_limit_up': is_limit_up,
            'time': datetime.datetime.now().isoformat()
        }

        try:
            h5_write_queue.put({
                'code': code,
                'date': today,
                'df': df_full,
                'indicators': indicators
            }, timeout=0.1)
        except queue.Full:
            with performance_lock:
                performance_stats['h5_queue_full_count'] += 1

        # 更新缓存（包含昨收盘价）
        with stock_data_cache_lock:
            stock_data_cache[code] = {
                'price': price,
                'vpr': vpr,
                'z_bias': z_bias,
                'intent': intent,
                'is_resonance': is_resonance,
                'signal_level': signal_level,
                'pre_close': pre_close,  # 缓存昨收盘价
                'time': datetime.datetime.now()
            }

        result = {
            'stock_data': stock_data,
            'is_resonance': is_resonance,
            'is_warning': is_warning,
            'is_strong': is_strong,
            'is_positive': is_positive,
            'signal_level': signal_level,
            'signal_score': score,
            'z_bias': z_bias,
            'intent': intent,
            'count': count,
            'is_new_alert': is_new_alert,
            'vpr': vpr,
            'confidence': confidence,
            'is_limit_down': is_limit_down,
            'is_limit_up': is_limit_up,
            'code': code
        }

        return result

    except Exception as e:
        if CONFIG['debug_mode']:
            print(f"{Fore.RED}❌ 处理股票异常 {code}: {e}")
        return None

# ==================== 扫描函数 ====================

def scan_stocks(codes):
    """扫描股票列表"""
    global scan_in_progress, scan_count

    with scan_lock:
        if scan_in_progress:
            return None
        scan_in_progress = True

    try:
        results = {
            'resonance': [],
            'warning': [],
            'strong': [],
            'positive': [],
            'normal': [],
            'top_signals': [],
            'combinations': []
        }

        api = get_best_server()
        if not api:
            print(f"{Fore.RED}❌ 无法连接服务器")
            return results

        for code in codes:
            result = process_stock(code, api, scan_count)
            if result:
                if result['is_resonance']:
                    results['resonance'].append(result)
                elif result['is_warning']:
                    results['warning'].append(result)
                elif result['is_strong']:
                    results['strong'].append(result)
                elif result['is_positive']:
                    results['positive'].append(result)
                else:
                    results['normal'].append(result)

                if result.get('signal_score', 0) >= CONFIG['min_display_score']:
                    results['top_signals'].append(result)

        api.disconnect()

        results['top_signals'].sort(key=lambda x: x['signal_score'], reverse=True)
        results['resonance'].sort(key=lambda x: x['signal_score'], reverse=True)
        results['strong'].sort(key=lambda x: x['signal_score'], reverse=True)
        results['positive'].sort(key=lambda x: x['signal_score'], reverse=True)
        results['warning'].sort(key=lambda x: x['signal_score'])

        for result in results['resonance'] + results['strong'] + results['warning'] + results['positive']:
            if result and 'stock_data' in result:
                combinations = check_signal_combinations(
                    result['code'],
                    result['stock_data'],
                    result,
                    results['top_signals'],
                    scan_count
                )
                results['combinations'].extend(combinations)

        unique_combos = {}
        for combo in results['combinations']:
            key = f"{combo['code']}_{combo['combo_key']}"
            if key not in unique_combos:
                unique_combos[key] = combo

        confidence_order = {'⭐⭐⭐⭐⭐⭐': 6, '⭐⭐⭐⭐⭐': 5, '⭐⭐⭐⭐': 4, '⭐⭐⭐': 3}
        results['combinations'] = sorted(
            unique_combos.values(),
            key=lambda x: (-confidence_order.get(x['confidence'], 0), x['code'])
        )[:SIGNAL_COMBINATION_CONFIG['display_top']]

        return results

    finally:
        with scan_lock:
            scan_in_progress = False

# ==================== HDF5写入线程 ====================

def h5_writer():
    """HDF5写入线程"""
    while True:
        try:
            item = h5_write_queue.get(timeout=1)
            if item == "STOP":
                break
            h5_save_data(item['code'], item['date'], item['df'], item['indicators'])
            h5_write_queue.task_done()
        except queue.Empty:
            continue
        except:
            try:
                h5_write_queue.task_done()
            except:
                pass

# ==================== 主程序 ====================

if __name__ == '__main__':
    if not os.path.exists('stocks.txt'):
        print(f"{Fore.RED}❌ 错误: 未找到 stocks.txt")
        print(f"{Fore.YELLOW}请创建stocks.txt文件，每行一个股票代码")
        exit()

    codes = []
    try:
        with open('stocks.txt', 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if line:
                    match = re.search(r'(\d{6})', line)
                    if match:
                        codes.append(match.group(1).zfill(6))
    except Exception as e:
        print(f"{Fore.RED}❌ 读取stocks.txt失败: {e}")
        exit()

    if not codes:
        print(f"{Fore.RED}❌ stocks.txt中没有股票代码")
        exit()

    DisplayManager.print_header()
    print(f"{Fore.GREEN}监控股票: {len(codes)}只")
    print(f"{Fore.GREEN}股票名称库: {len(STOCK_NAME_DB)}只")
    print(f"{Fore.YELLOW}启动时间: {datetime.datetime.now().strftime('%H:%M:%S')}")
    print(f"{Fore.CYAN}扫描间隔: {CONFIG['scan_interval']}秒")
    print(f"{Fore.CYAN}终端宽度: {DISPLAY_CONFIG['terminal_width']}列")
    print(f"{Fore.CYAN}显示模式: {'紧凑' if DISPLAY_CONFIG['compact_mode'] else '标准'}")
    print(f"{Fore.RED}⚠️ 已启用扫描并发防护")
    print(f"{Fore.RED}🔥 主力信心指标(VPR)已启用")
    print(f"{Fore.GREEN}✅ 涨跌停精准识别(增强) | 共振连续5日榜(修复) | 👑强庄标记 | 智能对齐")
    print(f"{Fore.GREEN}✅ 博弈比负偏取整修复 | 委比分语义优化")
    print(f"{Fore.GREEN}✅ 量额比逻辑修正：当前价/成交均价 | 红色=溢价拉升 | 绿色=折价压价")
    print(f"{Fore.GREEN}✅ 共振榜意图修复：显示当前共振意图")
    print(f"{Fore.GREEN}✅ 涨跌停增强：昨收盘价4层获取 | 板块差异识别 | 互斥判断 | 降级识别")
    print(f"{Fore.GREEN}✅ 连续天数修复：基于实际连续日期计算，频率与连续天数一致")

    print(f"{Fore.GREEN}✅ 多信号组合策略已加载 {len(SIGNAL_COMBINATION_CONFIG['combinations'])} 组")
    for key, combo in SIGNAL_COMBINATION_CONFIG['combinations'].items():
        print(f"{Fore.CYAN}   - {combo['name']}: {combo['confidence']} {combo['strategy']}")

    init_session_state()

    scan_count = SESSION_STATE['scan_count']
    resonance_total = SESSION_STATE['resonance_total']

    h5_thread = threading.Thread(target=h5_writer, daemon=True)
    h5_thread.start()

    try:
        while True:
            now = datetime.datetime.now()
            current_time = (now.hour, now.minute)

            is_trading = (
                    (9, 25) <= current_time <= (11, 31) or
                    (13, 0) <= current_time <= (15, 5)
            )

            if is_trading:
                scan_count += 1

                if scan_count == 1 or scan_count % 10 == 0:
                    DisplayManager.print_scan_info(scan_count, now.strftime('%H:%M:%S'))

                start_time = time.time()
                print(f"{Fore.CYAN}🚶 开始扫描 {len(codes)}只股票...")

                results = scan_stocks(codes)

                if results:
                    elapsed = time.time() - start_time

                    stats = {
                        'resonance': len(results['resonance']),
                        'warning': len(results['warning']),
                        'strong': len(results['strong']),
                        'positive': len(results['positive']),
                        'normal': len(results['normal'])
                    }

                    resonance_total += stats['resonance']

                    if scan_count % 5 == 0:
                        SESSION_STATE['scan_count'] = scan_count
                        SESSION_STATE['resonance_total'] = resonance_total
                        save_session_state()

                    if results['resonance']:
                        DisplayManager.print_resonance_header()
                        for r in results['resonance'][:10]:
                            DisplayManager.print_stock_info(r['stock_data'], 'resonance')

                    if results['warning']:
                        DisplayManager.print_warning_header()
                        for r in results['warning'][:10]:
                            DisplayManager.print_stock_info(r['stock_data'], 'warning')

                    if results['strong']:
                        DisplayManager.print_strong_header()
                        for r in results['strong'][:10]:
                            DisplayManager.print_stock_info(r['stock_data'], 'strong')

                    if results['top_signals']:
                        DisplayManager.print_top_signals_header()
                        for i, r in enumerate(results['top_signals'][:10], 1):
                            DisplayManager.print_stock_info(r['stock_data'], 'top', i)

                    if results['combinations']:
                        DisplayManager.print_combination_signals(results['combinations'])

                    if scan_count % 5 == 0:
                        DisplayManager.print_resonance_streak_board()

                    DisplayManager.print_stats(stats, warning_stats, elapsed, scan_count)

                time.sleep(CONFIG['scan_interval'])

            elif now.hour == 15 and now.minute >= 6:
                print(f"\n{Fore.YELLOW}🏁 交易结束")
                print(f"{Fore.CYAN}今日扫描: {scan_count}轮")
                print(f"{Fore.RED}核心共振: {resonance_total}次")
                print(f"{Fore.MAGENTA}累计报警: {warning_stats['total_warnings']}次")

                DisplayManager.print_resonance_streak_board()
                break

            else:
                time.sleep(30)

    except KeyboardInterrupt:
        print(f"\n{Fore.YELLOW}⚠️ 用户中断")
    except Exception as e:
        print(f"\n{Fore.RED}❌ 程序异常: {e}")
    finally:
        SESSION_STATE['scan_count'] = scan_count
        SESSION_STATE['resonance_total'] = resonance_total
        save_session_state()

        h5_write_queue.put("STOP")
        h5_thread.join(timeout=2)

        with resonance_history_lock:
            streak_3plus = len([d for d in resonance_streak.values() if d >= 3])
            streak_5 = len([d for d in resonance_streak.values() if d >= 5])

        with combination_history_lock:
            today = datetime.date.today().strftime('%Y-%m-%d')
            combo_today = len([v for v in combination_history.values()
                               if v.get('time', '').startswith(today)])

        print(f"{Fore.GREEN}✅ 会话已保存 | 今日扫描 {scan_count} 轮 | 共振 {resonance_total} 次")
        print(f"{Fore.GREEN}✅ 连续共振股: {len([d for d in resonance_streak.values() if d > 0])} 只")
        print(f"{Fore.GREEN}✅ 强庄股(3天+): {streak_3plus} 只 | 龙头股(5天): {streak_5} 只")
        print(f"{Fore.GREEN}✅ 今日组合信号: {combo_today} 个")
        print(f"{Fore.GREEN}✅ 量额比逻辑已修正：当前价/成交均价 | 红色=拉升 | 绿色=压价")
        print(f"{Fore.GREEN}✅ 共振连续榜意图已修复：显示当前共振意图")
        print(f"{Fore.GREEN}✅ 涨跌停判断已增强：昨收盘价4层获取 | 板块差异 | 互斥判断")
        print(f"{Fore.GREEN}✅ 连续天数已修复：基于实际连续日期计算")
        print(f"{Fore.GREEN}✅ 程序退出")